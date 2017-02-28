import column_name_formatter as cf
import ntpath
import openpyxl
import openpyxl.utils

from PyQt4.QtCore import QObject, pyqtSignal


class Value:
    def __init__(self, raw, format):
        self.raw = raw
        self.formatting = format

        self.formatted = format_value(raw, format)


class CellRef:
    """
    Helper class to encapsulate a cells location in the sheet, as well as its index into the value array
    """

    def __init__(self, idx, sheet_row, sheet_col):
        self.idx = idx
        self.sheet_row = sheet_row
        self.sheet_column = sheet_col


class ExcelBook(QObject):
    started = pyqtSignal()
    finished = pyqtSignal()

    progress = pyqtSignal('QString', int)

    def __init__(self):
        QObject.__init__(self)
        self.sheets = {}
        self.years = []
        self.name = ""
        self.loaded = False
        self._book = None

    def load(self, path):
        self.started.emit()
        out = False

        try:
            self.progress.emit("Loading book...", 0)
            self._book = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            pass

        if self._book:
            self.years = []
            self.name = ntpath.basename(path)

            names = self._book.sheetnames
            for n in names:
                try:
                    year = int(n)
                    self.years.append(str(year))
                except ValueError:
                    pass

            if self.years:
                ctr = 0

                self.sheets = {}
                for year in self.years:
                    sheet = self._book.get_sheet_by_name(year)
                    self.sheets[year] = ExcelSheet(year, sheet)
                    ctr += 1
                    self.progress.emit("Sheet {} loaded".format(year), int(0.5 + (100.0 * ctr) / len(self.years)))

                out = True

        self.loaded = out
        self.finished.emit()

        return out

    def get_data(self, year, constituency, column_name):
        """
        Return a single data value from a constituency given the year and column header

        Args:
            year: which year
            constituency: which constituency
            column_name: which data type

        Returns:
            Formatted string with the value
        """
        if year in self.sheets:
            return self.sheets[year].get_value(constituency, column_name)
        else:
            return None

    def get_constituency_data(self, constituency):
        """
        Get all data from a constituency for all years

        Args:
            constituency: name of the constituency
a
        Returns:
            Dictionary of years mapped to a dictionary of column headers mapped to the formatted values
        """
        return {year: self.sheets[year].get_constituency_data(constituency) for year in self.sheets}

    def get_column_header_data(self, column_header):
        """
        Get all data from a column header for all years

        Args:
            column_header: name of the column

        Returns:
            Dictionary of years mapped to a dictionary of constituencies mapped to the formatted values
        """
        return {year: self.sheets[year].get_column_header_data(column_header) for year in self.sheets}

    def get_headers(self):
        """
        Returns the headers for each year

        Returns:
            Dict of years mapped to a list of headers
        """
        return {year: self.sheets[year].column_headers for year in self.sheets}

    def get_constituencies(self):
        """
        Returns the constituencies for each year

        Returns:
            Dict of years mapped to a list of constituencies
        """
        return {year: self.sheets[year].constituencies for year in self.sheets}

    def get_empty_cells(self):
        """
        Returns the cells that contain no data

        Returns:
            Dict of years mapped to constituencies mapped to a list headers

        ex: empty_cells = doc.get_empty_cells()
            print(empty_cells["2012"])
                "Angus": [empty_column1, empty_column2],
                "another place": [empty_column1, empty_column3]
        """
        return {year: self.sheets[year].empty_cells for year in self.sheets if self.sheets[year].empty_cells}

    def get_dodgy_cells(self):
        """
        Returns the cells that were flagged as having data that couldn't be formatted for any reason

        Returns:
            Dict of years mapped to constituencies mapped to a list headers

        ex: dodgy_cells = doc.get_dodgy_cells()
            print(dodgy_cells["2012"])
                "Angus": [dodgy_column1, dodgy_column2],
                "another place": [dodgy_column1, dodgy_column3]
        """
        return {year: self.sheets[year].dodgy_cells for year in self.sheets if self.sheets[year].dodgy_cells}

    def column_exists(self, column):
        if self.loaded and self.years:
            sheet = self.sheets[self.years[0]]

            return sheet.column_exists(column)
        else:
            return False


class ExcelSheet:
    """
    Encapsulates an excel sheet and allows access to a constituencies data

    The excel sheet reference itself is not stored and instead all data is pulled from the sheet and is retrieved as a
    formatted string rather than accessing the sheet directly in calls. This should speed up the app since accessing the
    excel sheet should be minimised.

    All constituencies and column headers are mapped to integers starting at 0. The formatted data is then stored in a
    2D array and so calls to get(constituency, column_name) are mapped to integers and then pulled from the array. This
    prevents any excess data duplication

    Public calls use the raw column name, internal mappings use standard names
    """

    MAX_STARTING_COLUMN = 8
    MAX_STARTING_ROW = 16

    def __init__(self, name, sheet):
        """
        Read all data form the sheet and stores it in a 2D array of formatted strings
        """

        self.name = name
        self.constituencies = []
        self.column_headers = []  # raw names for public use
        self.empty_cells = {}  # cells with no data in them, uses raw names
        self.dodgy_cells = {}  # cells that aren't ints or floats, need to be checked for errors manually, raw names

        self._constituency_map = {}  # map of constituent names to CellRef
        self._column_header_map = {}  # column names to CellRef, uses standard name for mappings
        self._format_map = {}  # map of standard column names to the style of all cells in the column
        self._data_limits = self._set_data_limits(sheet)
        self._values = []  # 2D array, mapped by constituent and column names through the above maps

        if self._data_limits:
            self._build_constituency_map(sheet)
            self._build_column_header_maps(sheet)

            self._set_values(sheet)

    def get_value(self, constituency, column_header):
        """
        Return a formatted string representing the excel cell for a constituency and column header

        Args:
            constituency: name of the constituency
            column_header: name of the column header, name is not formatted

        Returns:
            Formatted string representing the excel cell for a constituency and column header
        """
        standard_name = cf.fmt(column_header)

        row_idx = self._constituency_map[constituency].idx
        column_idx = self._column_header_map[standard_name].idx

        return self._values[row_idx][column_idx]

    def get_constituency_data(self, constituency):
        """
        Get all data from a constituency

        Args:
            constituency: name of the constituency

        Returns:
            Dictionary of column headers mapped to the formatted values
        """
        return {column_header: self.get_value(constituency, column_header) for column_header in self.column_headers}

    def get_column_header_data(self, column_header):
        """
        Get all data from a column header

        Args:
            column_header: name of the column

        Returns:
            Dictionary of constituencies mapped to the formatted values
        """
        return {constituency: self.get_value(constituency, column_header) for constituency in self.constituencies}

    def _set_data_limits(self, sheet):
        """
        Returns a map representing the top-left corner of the data and the bottom right corner

        The top-left is where the 'Constituency' or 'Local Authority' cell is, with all cells below being the
        constituents and all cells to the right being the column headings. This must appear in the first 8 columns and
        first 16 rows of the excel sheet.

        The bottom-right is where the data ends. The column number is found from max_column. The max rows may be
        different from max_row due to extra notes at the bottom of the sheet

        Args:
            sheet: reference to worksheet

        Returns:
            Tuple with row and column of location of the 'Constituency' cell.
            Values are None if no cell is found
        """
        limits = None

        max_col = min(self.MAX_STARTING_COLUMN, sheet.max_column)
        max_row = min(self.MAX_STARTING_ROW, sheet.max_row)

        for column in range(1, max_col + 1):
            for row in range(1, max_row + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if cell_value and str(cell_value).strip() in ["Constituency", "Local Authority"]:
                    limits = {"start-row": row, "start-column": column}
                    break

            if limits:
                break

        if limits:
            limits["end-column"] = sheet.max_column
            for row in range(sheet.max_row, limits["start-row"], -1):
                constituent = sheet.cell(row=row, column=limits["start-column"]).value
                if constituent and str(constituent).strip() in ["Total Clients", "All Constituents"]:
                    limits["end-row"] = row

            if "end-row" not in limits:
                limits["end-row"] = sheet.max_row

        return limits

    def _build_constituency_map(self, sheet):
        """
        Generate a mapping between constituents and integers for array indexing

        The order in which they appear define the indexes

        Args:
            sheet: reference to worksheet
        """
        column = self._data_limits["start-column"]
        start_row = self._data_limits["start-row"] + 1
        end_row = self._data_limits["end-row"] + 1

        idx = 0
        for row in range(start_row, end_row):
            constituent = sheet.cell(row=row, column=column).value

            self.constituencies.append(constituent)
            self._constituency_map[constituent] = CellRef(idx=idx, sheet_row=row, sheet_col=column)
            idx += 1

    def _build_column_header_maps(self, sheet):
        """
        Generate a mapping between column headers and integers for array indexing. Also generates the value format map

        The order in which they appear define the indexes

        Args:
            sheet: reference to worksheet
        """
        row_idx = self._data_limits["start-row"]
        start_column = self._data_limits["start-column"] + 1
        end_column = self._data_limits["end-column"] + 1

        idx = 0
        for column_idx in range(start_column, end_column):
            raw_column_header = sheet.cell(row=row_idx, column=column_idx).value
            if raw_column_header:
                standard_name = cf.fmt(raw_column_header)
                column_format = sheet.cell(row=row_idx + 1, column=column_idx).number_format

                self.column_headers.append(raw_column_header)
                self._column_header_map[standard_name] = CellRef(idx=idx, sheet_row=row_idx, sheet_col=column_idx)
                self._format_map[standard_name] = column_format
                idx += 1

    def _set_values(self, sheet):
        """
        Build a 2D grid of formatted strings that represent the excel data to be put into the word docs

        Args:
            sheet: reference to excel worksheet
        """
        self._values = [[None] * len(self.column_headers) for _ in range(0, len(self.constituencies))]

        for constituency in self.constituencies:
            cell_ref = self._constituency_map[constituency]
            row_idx = cell_ref.idx
            sheet_row = cell_ref.sheet_row

            for raw_header in self.column_headers:
                standard_name = cf.fmt(raw_header)

                col_idx = self._column_header_map[standard_name].idx
                sheet_column = self._column_header_map[standard_name].sheet_column

                cell_value = sheet.cell(row=sheet_row, column=sheet_column).value

                if cell_value:
                    try:
                        value = Value(cell_value, self._format_map[standard_name])
                        self._values[row_idx][col_idx] = value
                    except TypeError:
                        # Most cells are either ints, floats or empty
                        # Some seem to be 1-length strings though and so may need to be looked at
                        self._values[row_idx][col_idx] = ""
                        if constituency not in self.dodgy_cells:
                            self.dodgy_cells[constituency] = []

                        self.dodgy_cells[constituency].append(raw_header)
                else:
                    # saving empty columns to display later
                    self._values[row_idx][col_idx] = ""
                    if constituency not in self.empty_cells:
                        self.empty_cells[constituency] = []

                    self.empty_cells[constituency].append(raw_header)

    def column_exists(self, column):
        return cf.fmt(column) in self._column_header_map


def format_value(raw, formatting):
    """
    Return a cell value with correct formatting

    Args:
        raw: cell value
        formatting: excel numerical format

    Returns:
        Value formatted to match the excel sheet
    """
    formatted = raw
    if "%" in formatting:
        formatted *= 100

    if "0.00" in formatting:
        formatted = "%.2f" % formatted
    elif "0.0" in formatting:
        formatted = "%.1f" % formatted
    else:
        formatted = "%.0f" % formatted

    if "£" in formatting:
        formatted = "£" + formatted
    elif "%" in formatting:
        formatted += "%"

    return formatted
